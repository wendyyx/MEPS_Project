# %% 0. Imports
import re
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns
import statsmodels.formula.api as smf
import statsmodels.api as sm
from scipy import stats
from statsmodels.stats.diagnostic import het_breuschpagan
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# %% ─
DATA_DIR   = Path(r"C:\Users\wenzp\Desktop\WISC\Summer\AAE 718-DS\Week 4\Project 4\meps_project\data\raw")
CLEAN_DIR  = Path(r"C:\Users\wenzp\Desktop\WISC\Summer\AAE 718-DS\Week 4\Project 4\meps_project\data\clean")
OUTPUT_DIR = Path(r"C:\Users\wenzp\Desktop\WISC\Summer\AAE 718-DS\Week 4\Project 4\meps_project\outputs")

for d in [CLEAN_DIR, OUTPUT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

DAT_FILE = DATA_DIR / "h243.dat"
SAS_FILE = DATA_DIR / "h243su.txt"

# %% 
sas_text = SAS_FILE.read_text(encoding="latin-1", errors="replace")

PATTERN = re.compile(
    r'@\s*(\d+)\s+([A-Z][A-Z0-9_]*)\s+(\$?)(\d+)\.(\d*)',
    re.IGNORECASE
)

all_cols = []
for m in PATTERN.finditer(sas_text):
    start   = int(m.group(1)) - 1
    varname = m.group(2).upper()
    is_char = m.group(3) == "$"
    width   = int(m.group(4))
    all_cols.append({
        "name":    varname,
        "start":   start,
        "end":     start + width,
        "is_char": is_char,
    })

print(f"  Found {len(all_cols)} variables in format file.")

# %% Select variables & load data
KEEP_VARS = {
    "DUPERSID": "person_id",
    "INSCOV22": "insured_status",
    "RTHLTH53": "self_rated_health",
    "AGE22X":   "age",
    "SEX":      "female",
    "POVCAT22": "income_cat",
    "EDUCYR":   "educ_years",
    "RACETHX":  "race_eth",
    "MARRY22X": "married",
    "EMPST53":  "employed",
    "REGION22": "region",
}

selected = [c for c in all_cols if c["name"] in KEEP_VARS]
colspecs  = [(c["start"], c["end"]) for c in selected]
colnames  = [c["name"] for c in selected]

raw = pd.read_fwf(
    DAT_FILE,
    colspecs=colspecs,
    names=colnames,
    encoding="latin-1",
    dtype=str,
)
print(f"  Raw shape: {raw.shape[0]:,} rows × {raw.shape[1]} columns")

raw = raw.rename(columns=KEEP_VARS)
for col in raw.columns:
    if col != "person_id":
        raw[col] = pd.to_numeric(raw[col], errors="coerce")

# %% Data cleaning
df    = raw.copy()
n_raw = len(df)

MISSING_CODES = {-1, -7, -8, -9, -15}
for col in df.columns:
    if col == "person_id":
        continue
    df[col] = df[col].where(~df[col].isin(MISSING_CODES), other=np.nan)

# Outcome
df["insured"]     = (df["insured_status"] != 3).astype(float)
df["private_ins"] = (df["insured_status"] == 1).astype(float)
df.loc[df["insured_status"].isna(), ["insured", "private_ins"]] = np.nan

# Self-rated health: reverse so 5=Excellent, 1=Poor
df["health_score"] = 6 - df["self_rated_health"]
df["poor_health"]  = (df["self_rated_health"] >= 4).astype(float)
df.loc[df["self_rated_health"].isna(), ["health_score", "poor_health"]] = np.nan

HEALTH_LABELS = {1: "Poor", 2: "Fair", 3: "Good", 4: "Very Good", 5: "Excellent"}

# Dummies
df["female"]   = (df["female"]   == 2).astype(float)
df["married"]  = (df["married"]  == 1).astype(float)
df["employed"] = (df["employed"] == 1).astype(float)

df["hispanic"] = (df["race_eth"] == 1).astype(float)
df["black"]    = (df["race_eth"] == 3).astype(float)
df["asian"]    = (df["race_eth"] == 4).astype(float)
for c in ["hispanic", "black", "asian"]:
    df.loc[df["race_eth"].isna(), c] = np.nan

df["age_sq"] = df["age"] ** 2

# Sample restriction
df = df[(df["age"] >= 18) & (df["age"] <= 64)]
print(f"  After age 18–64 restriction : {len(df):,}")

MODEL_VARS = [
    "insured", "private_ins",
    "health_score", "poor_health",
    "age", "age_sq", "female",
    "income_cat", "educ_years",
    "hispanic", "black", "asian",
    "married", "employed", "region",
]
MODEL_VARS = [v for v in MODEL_VARS if v in df.columns]
analytic   = df[MODEL_VARS].dropna()
n_final    = len(analytic)

print(f"  Final analytic sample : {n_final:,}")
print(f"  Total excluded        : {n_raw - n_final:,} ({(n_raw-n_final)/n_raw*100:.1f}%)")

analytic.to_csv(CLEAN_DIR / "meps_analytic.csv", index=False)

# %% Descriptive statistics
desc = analytic.describe().T[["count", "mean", "std", "min", "max"]].round(3)
print(desc.to_string())
desc.to_csv(OUTPUT_DIR / "table1_descriptive.csv")

# %% Coverage rate by health score
fig1, ax1 = plt.subplots(figsize=(8, 5))
grp      = analytic.groupby("health_score")["insured"].mean() * 100
ns       = analytic.groupby("health_score")["insured"].count()
xlabels  = [HEALTH_LABELS.get(int(k), k) for k in grp.index]

bars = ax1.bar(xlabels, grp.values,
               color=sns.color_palette("Blues_d", len(grp)),
               edgecolor="white", width=0.6)
for bar, n in zip(bars, ns.values):
    ax1.text(bar.get_x() + bar.get_width() / 2,
             bar.get_height() + 0.6,
             f"n={n:,}", ha="center", va="bottom", fontsize=8, color="grey")

ax1.set_xlabel("Self-Rated Health Status", fontsize=11)
ax1.set_ylabel("Insurance Coverage Rate (%)", fontsize=11)
ax1.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.0f%%"))
ax1.set_ylim(0, 108)
fig1.tight_layout()
fig1.savefig(OUTPUT_DIR / "fig1_coverage_by_health.png", dpi=150)
plt.show()

# %% Health score distribution
fig2, ax2 = plt.subplots(figsize=(8, 5))
hcounts = analytic["health_score"].value_counts().sort_index()
ax2.bar([HEALTH_LABELS.get(int(k), k) for k in hcounts.index],
        hcounts.values,
        color=sns.color_palette("Greens_d", len(hcounts)),
        edgecolor="white", width=0.6)
ax2.set_xlabel("Self-Rated Health Status", fontsize=11)
ax2.set_ylabel("Number of Respondents", fontsize=11)
ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x):,}"))
fig2.tight_layout()
fig2.savefig(OUTPUT_DIR / "fig2_health_distribution.png", dpi=150)
plt.show()

# %% Regression models
demo  = [v for v in ["age", "age_sq", "female"] if v in analytic.columns]
race  = [v for v in ["hispanic", "black", "asian"] if v in analytic.columns]
socio = [v for v in ["income_cat", "educ_years"] if v in analytic.columns]
lab   = [v for v in ["married", "employed"] if v in analytic.columns]
reg   = [v for v in ["region"] if v in analytic.columns]

SPECS = {
    "(1)": ["health_score"],
    "(2)": ["health_score"] + demo + race,
    "(3)": ["health_score"] + demo + race + socio,
    "(4)": ["health_score"] + demo + race + socio + lab + reg,
}

fitted = {}
for col_label, preds in SPECS.items():
    preds              = [p for p in preds if p in analytic.columns]
    formula            = "insured ~ " + " + ".join(preds)
    fitted[col_label]  = smf.ols(formula, data=analytic).fit(cov_type="HC3")
    coef = fitted[col_label].params.get("health_score", np.nan)
    pval = fitted[col_label].pvalues.get("health_score", np.nan)
    sig  = "***" if pval < 0.01 else ("**" if pval < 0.05 else ("*" if pval < 0.1 else ""))
    print(f"  {col_label}  β(health_score)={coef:+.4f}{sig}  "
          f"R²={fitted[col_label].rsquared:.4f}  N={int(fitted[col_label].nobs):,}")

# %% Export regression table to Excel
VAR_LABELS = [
    ("health_score", "Health score (1=Poor, 5=Excellent)"),
    ("age",          "Age"),
    ("age_sq",       "Age squared"),
    ("female",       "Female (=1)"),
    ("hispanic",     "Hispanic (ref: NH White)"),
    ("black",        "NH Black (ref: NH White)"),
    ("asian",        "NH Asian (ref: NH White)"),
    ("income_cat",   "Income category (1–5)"),
    ("educ_years",   "Years of education"),
    ("married",      "Married (=1)"),
    ("employed",     "Employed (=1)"),
    ("region",       "Census region (1–4)"),
    ("Intercept",    "Constant"),
]

MODEL_COLS = list(SPECS.keys())
CONTROLS   = {
    "(1)": ["—", "—", "—", "—"],
    "(2)": ["✓", "—", "—", "—"],
    "(3)": ["✓", "✓", "—", "—"],
    "(4)": ["✓", "✓", "✓", "✓"],
}
CTRL_LABELS = [
    "Demographics (age, sex)",
    "Race / ethnicity",
    "Socioeconomic (income, educ.)",
    "Labour market & region",
]

# Style helpers
wb    = Workbook()
ws    = wb.active
ws.title = "Regression Results"

BLACK    = "000000"
WHITE    = "FFFFFF"
GRAY_HD  = "F2F2F2"
BLUE_KEY = "1F4E79"
thick    = Side(style="medium", color=BLACK)

def wcell(ws, row, col, value,
          bold=False, italic=False, fc=BLACK,
          bg=None, align="center", size=11,
          top=False, bottom=False):
    c            = ws.cell(row=row, column=col, value=value)
    c.border     = Border(
        top    = thick if top    else Side(style=None),
        bottom = thick if bottom else Side(style=None),
    )
    c.font       = Font(name="Times New Roman", bold=bold,
                        italic=italic, color=fc, size=size)
    c.alignment  = Alignment(horizontal=align, vertical="center",
                             wrap_text=False)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c

# Column widths
ws.column_dimensions["A"].width = 38
for ci in range(2, 6):
    ws.column_dimensions[get_column_letter(ci)].width = 14

# Row 1: table title
ws.merge_cells("A1:E1")
c            = ws["A1"]
c.value      = "Table 2. Effect of Self-Rated Health on Insurance Coverage"
c.font       = Font(name="Times New Roman", bold=True, size=12)
c.alignment  = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 20

# Row 2: outcome label + model column headers (top border)
ws.row_dimensions[2].height = 18
wcell(ws, 2, 1, "Outcome: P(Insured)", bold=True, align="left", top=True)
for i, label in enumerate(MODEL_COLS, start=2):
    wcell(ws, 2, i, label, bold=True, align="center", top=True)

# Row 3: thin separator
ws.row_dimensions[3].height = 4
for col in range(1, 6):
    ws.cell(row=3, column=col).border = Border(
        bottom=Side(style="thin", color=BLACK)
    )

# Coefficient and SE rows
cur_row = 4
for var_key, var_label in VAR_LABELS:
    is_key = (var_key == "health_score")
    fc     = BLUE_KEY if is_key else BLACK

    # Coefficient row
    wcell(ws, cur_row, 1, var_label, bold=is_key, fc=fc,
          align="left", size=10)
    for i, mc in enumerate(MODEL_COLS, start=2):
        m = fitted[mc]
        if var_key in m.params:
            coef = m.params[var_key]
            pval = m.pvalues[var_key]
            sig  = ("***" if pval < 0.01 else
                    "**"  if pval < 0.05 else
                    "*"   if pval < 0.1  else "")
            val  = f"{coef:+.4f}{sig}"
        else:
            val = ""
        wcell(ws, cur_row, i, val, bold=is_key, fc=fc,
              align="center", size=10)
    ws.row_dimensions[cur_row].height = 14
    cur_row += 1

    # SE row (italic, grey, parentheses)
    wcell(ws, cur_row, 1, "", align="left", size=10)
    for i, mc in enumerate(MODEL_COLS, start=2):
        m   = fitted[mc]
        val = f"({m.bse[var_key]:.4f})" if var_key in m.bse else ""
        wcell(ws, cur_row, i, val, italic=True,
              fc="595959", align="center", size=10)
    ws.row_dimensions[cur_row].height = 13
    cur_row += 1

# Thin separator before controls
ws.row_dimensions[cur_row].height = 4
for col in range(1, 6):
    ws.cell(row=cur_row, column=col).border = Border(
        top=Side(style="thin", color=BLACK)
    )
cur_row += 1

# Control variable indicators
for ctrl_label, ctrl_vals in zip(
    CTRL_LABELS,
    zip(*[CONTROLS[mc] for mc in MODEL_COLS])
):
    wcell(ws, cur_row, 1, ctrl_label, align="left",
          size=10, bg=GRAY_HD)
    for i, val in enumerate(ctrl_vals, start=2):
        wcell(ws, cur_row, i, val, align="center",
              size=10, bg=GRAY_HD)
    ws.row_dimensions[cur_row].height = 14
    cur_row += 1

# Thin separator before N / R²
ws.row_dimensions[cur_row].height = 4
for col in range(1, 6):
    ws.cell(row=cur_row, column=col).border = Border(
        top=Side(style="thin", color=BLACK)
    )
cur_row += 1

# N and R² rows (bottom border on R²)
for stat_label, extractor in [
    ("N",  lambda m: int(m.nobs)),
    ("R²", lambda m: round(m.rsquared, 4)),
]:
    is_last = (stat_label == "R²")
    wcell(ws, cur_row, 1, stat_label, bold=True, align="left",
          size=10, bottom=is_last)
    for i, mc in enumerate(MODEL_COLS, start=2):
        wcell(ws, cur_row, i, extractor(fitted[mc]),
              bold=True, align="center", size=10, bottom=is_last)
    ws.row_dimensions[cur_row].height = 14
    cur_row += 1

# Notes row
ws.row_dimensions[cur_row].height = 6
cur_row += 1
ws.merge_cells(f"A{cur_row}:E{cur_row}")
c            = ws.cell(row=cur_row, column=1,
                       value=("Notes: HC3 heteroskedasticity-robust standard errors "
                              "in parentheses. *** p<0.01, ** p<0.05, * p<0.10. "
                              "Reference group: Non-Hispanic White. "
                              "Data: MEPS 2022, adults aged 18–64."))
c.font       = Font(name="Times New Roman", italic=True, size=9, color="595959")
c.alignment  = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[cur_row].height = 28

ws.freeze_panes = "A4"

out_path = OUTPUT_DIR / "Table2_regression_results.xlsx"
wb.save(out_path)
print(f"\n  Saved → {out_path}")
